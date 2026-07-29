import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from io import BytesIO


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Excel Report Generator",
    page_icon="📊",
    layout="wide"
)


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.header("📊 Excel Report Generator")

    st.write(
        "Automated data analysis and Excel reporting tool."
    )

    st.divider()

    st.subheader("Features")

    st.write("✓ CSV data upload")
    st.write("✓ Data cleaning")
    st.write("✓ Summary statistics")
    st.write("✓ Sales analysis")
    st.write("✓ Data visualisation")
    st.write("✓ Excel report generation")

    st.divider()

    st.subheader("Technologies")

    st.write("🐍 Python")
    st.write("🐼 Pandas")
    st.write("📈 Matplotlib")
    st.write("📗 OpenPyXL")
    st.write("🖥️ Streamlit")

    st.divider()

    st.caption(
        "Developed as an internship project."
    )


# ============================================================
# MAIN HEADER
# ============================================================

st.title("📊 Excel Report Generator")

st.markdown(
    """
    ### Automated Data Analysis & Reporting

    Upload a CSV dataset to analyse your data, generate
    visualisations, and export a professionally structured
    Excel report.
    """
)

st.divider()


# ============================================================
# FILE UPLOAD
# ============================================================

st.subheader("📂 Upload Dataset")

uploaded_file = st.file_uploader(
    "Choose a CSV file",
    type=["csv"],
    help="Upload a CSV file containing your dataset."
)


# ============================================================
# MAIN APPLICATION
# ============================================================

if uploaded_file is not None:

    try:

        # ----------------------------------------------------
        # READ DATA
        # ----------------------------------------------------

        df = pd.read_csv(uploaded_file)

        st.success(
            f"Successfully loaded: {uploaded_file.name}"
        )

        # ----------------------------------------------------
        # DATA CLEANING
        # ----------------------------------------------------

        original_rows = len(df)

        # Remove completely empty rows
        df = df.dropna(how="all")

        # Remove duplicate rows
        df = df.drop_duplicates()

        removed_rows = original_rows - len(df)

        # ----------------------------------------------------
        # CALCULATE TOTAL SALES
        # ----------------------------------------------------

        if (
            "Quantity" in df.columns
            and "Unit_Price" in df.columns
        ):

            df["Total_Sales"] = (
                df["Quantity"] * df["Unit_Price"]
            )

        # ====================================================
        # DATASET OVERVIEW
        # ====================================================

        st.subheader("📋 Dataset Overview")

        col1, col2, col3, col4 = st.columns(4)

        with col1:

            st.metric(
                "Total Records",
                f"{len(df):,}"
            )

        with col2:

            if "Total_Sales" in df.columns:

                total_sales = df["Total_Sales"].sum()

                st.metric(
                    "Total Sales",
                    f"₹{total_sales:,.0f}"
                )

            else:

                st.metric(
                    "Total Sales",
                    "N/A"
                )

        with col3:

            if "Total_Sales" in df.columns:

                average_sales = df[
                    "Total_Sales"
                ].mean()

                st.metric(
                    "Average Sales",
                    f"₹{average_sales:,.0f}"
                )

            else:

                st.metric(
                    "Average Sales",
                    "N/A"
                )

        with col4:

            st.metric(
                "Total Columns",
                len(df.columns)
            )

        # ----------------------------------------------------
        # DATA CLEANING INFORMATION
        # ----------------------------------------------------

        if removed_rows > 0:

            st.info(
                f"{removed_rows} duplicate/empty rows "
                "were removed during data cleaning."
            )

        else:

            st.info(
                "No duplicate or completely empty rows found."
            )

        # ====================================================
        # DATA PREVIEW
        # ====================================================

        st.subheader("🔍 Data Preview")

        st.dataframe(
            df,
            use_container_width=True,
            height=300
        )

        # ====================================================
        # SUMMARY STATISTICS
        # ====================================================

        st.subheader("📈 Summary Statistics")

        numeric_columns = df.select_dtypes(
            include="number"
        ).columns

        if len(numeric_columns) > 0:

            summary = df[
                numeric_columns
            ].describe()

            st.dataframe(
                summary,
                use_container_width=True
            )

        else:

            st.warning(
                "No numerical columns were found."
            )

        # ====================================================
        # SALES ANALYSIS
        # ====================================================

        if "Total_Sales" in df.columns:

            st.divider()

            st.header("📊 Sales Analysis")

            # ------------------------------------------------
            # CATEGORY ANALYSIS
            # ------------------------------------------------

            if "Category" in df.columns:

                st.subheader("🛍️ Sales by Category")

                category_sales = (
                    df.groupby("Category")[
                        "Total_Sales"
                    ]
                    .sum()
                    .sort_values(
                        ascending=False
                    )
                )

                fig, ax = plt.subplots(
                    figsize=(9, 5)
                )

                category_sales.plot(
                    kind="bar",
                    ax=ax
                )

                ax.set_title(
                    "Sales by Category"
                )

                ax.set_xlabel(
                    "Category"
                )

                ax.set_ylabel(
                    "Total Sales"
                )

                plt.xticks(
                    rotation=45
                )

                plt.tight_layout()

                st.pyplot(fig)

                plt.close(fig)

            # ------------------------------------------------
            # REGION ANALYSIS
            # ------------------------------------------------

            if "Region" in df.columns:

                st.subheader("🌍 Sales by Region")

                region_sales = (
                    df.groupby("Region")[
                        "Total_Sales"
                    ]
                    .sum()
                    .sort_values(
                        ascending=False
                    )
                )

                fig, ax = plt.subplots(
                    figsize=(9, 5)
                )

                region_sales.plot(
                    kind="bar",
                    ax=ax
                )

                ax.set_title(
                    "Sales by Region"
                )

                ax.set_xlabel(
                    "Region"
                )

                ax.set_ylabel(
                    "Total Sales"
                )

                plt.xticks(
                    rotation=0
                )

                plt.tight_layout()

                st.pyplot(fig)

                plt.close(fig)

            # =================================================
            # TOP PRODUCTS
            # =================================================

            if "Product" in df.columns:

                st.subheader("🏆 Top Products")

                product_sales = (
                    df.groupby("Product")[
                        "Total_Sales"
                    ]
                    .sum()
                    .sort_values(
                        ascending=False
                    )
                )

                top_products = (
                    product_sales
                    .reset_index()
                )

                top_products.columns = [
                    "Product",
                    "Total Sales"
                ]

                st.dataframe(
                    top_products,
                    use_container_width=True
                )

        # ====================================================
        # EXCEL REPORT GENERATOR
        # ====================================================

        st.divider()

        st.header("📥 Excel Report")

        st.write(
            "Generate a formatted Excel workbook containing "
            "the analysed data and summary information."
        )

        if st.button(
            "📊 Generate Excel Report",
            use_container_width=True
        ):

            output = BytesIO()

            workbook = Workbook()

            # =================================================
            # SUMMARY SHEET
            # =================================================

            summary_sheet = workbook.active

            summary_sheet.title = "Summary"

            summary_sheet["A1"] = (
                "Excel Report Generator"
            )

            summary_sheet["A1"].font = Font(
                bold=True,
                size=18
            )

            summary_sheet["A3"] = (
                "Total Records"
            )

            summary_sheet["B3"] = len(df)

            summary_sheet["A4"] = (
                "Total Columns"
            )

            summary_sheet["B4"] = len(
                df.columns
            )

            if "Total_Sales" in df.columns:

                summary_sheet["A5"] = (
                    "Total Sales"
                )

                summary_sheet["B5"] = (
                    df["Total_Sales"].sum()
                )

                summary_sheet["A6"] = (
                    "Average Sales"
                )

                summary_sheet["B6"] = (
                    df["Total_Sales"].mean()
                )

            # =================================================
            # DATA SHEET
            # =================================================

            data_sheet = workbook.create_sheet(
                "Data"
            )

            # Headers
            for col_num, column_name in enumerate(
                df.columns,
                1
            ):

                cell = data_sheet.cell(
                    row=1,
                    column=col_num
                )

                cell.value = column_name

                cell.font = Font(
                    bold=True
                )

                cell.alignment = Alignment(
                    horizontal="center"
                )

            # Data
            for row_num, row in enumerate(
                df.itertuples(
                    index=False
                ),
                2
            ):

                for col_num, value in enumerate(
                    row,
                    1
                ):

                    data_sheet.cell(
                        row=row_num,
                        column=col_num,
                        value=value
                    )

            # =================================================
            # CATEGORY ANALYSIS SHEET
            # =================================================

            if (
                "Category" in df.columns
                and "Total_Sales" in df.columns
            ):

                category_sheet = (
                    workbook.create_sheet(
                        "Category Analysis"
                    )
                )

                category_sales = (
                    df.groupby("Category")[
                        "Total_Sales"
                    ]
                    .sum()
                    .reset_index()
                )

                category_sheet["A1"] = (
                    "Category"
                )

                category_sheet["B1"] = (
                    "Total Sales"
                )

                for cell in category_sheet[1]:

                    cell.font = Font(
                        bold=True
                    )

                    cell.alignment = Alignment(
                        horizontal="center"
                    )

                for row_num, row in enumerate(
                    category_sales.itertuples(
                        index=False
                    ),
                    2
                ):

                    category_sheet.cell(
                        row=row_num,
                        column=1,
                        value=row[0]
                    )

                    category_sheet.cell(
                        row=row_num,
                        column=2,
                        value=row[1]
                    )

                # ------------------------------------------------
                # EXCEL BAR CHART
                # ------------------------------------------------

                chart = BarChart()

                chart.title = (
                    "Sales by Category"
                )

                chart.y_axis.title = (
                    "Total Sales"
                )

                chart.x_axis.title = (
                    "Category"
                )

                data = Reference(
                    category_sheet,
                    min_col=2,
                    min_row=1,
                    max_row=(
                        category_sheet.max_row
                    )
                )

                categories = Reference(
                    category_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=(
                        category_sheet.max_row
                    )
                )

                chart.add_data(
                    data,
                    titles_from_data=True
                )

                chart.set_categories(
                    categories
                )

                category_sheet.add_chart(
                    chart,
                    "D2"
                )

            # =================================================
            # FORMAT ALL SHEETS
            # =================================================

            for sheet in workbook.worksheets:

                for column in sheet.columns:

                    max_length = 0

                    column_letter = (
                        column[0].column_letter
                    )

                    for cell in column:

                        if cell.value is not None:

                            max_length = max(
                                max_length,
                                len(
                                    str(
                                        cell.value
                                    )
                                )
                            )

                    sheet.column_dimensions[
                        column_letter
                    ].width = min(
                        max_length + 2,
                        40
                    )

                for row in sheet.iter_rows():

                    for cell in row:

                        cell.alignment = (
                            Alignment(
                                vertical="center"
                            )
                        )

            # =================================================
            # SAVE EXCEL FILE
            # =================================================

            workbook.save(output)

            output.seek(0)

            st.success(
                "✅ Excel report generated successfully!"
            )

            st.download_button(
                label="⬇️ Download Excel Report",
                data=output,
                file_name=(
                    "sales_analysis_report.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True
            )

    except Exception as e:

        st.error(
            f"❌ An error occurred: {e}"
        )


# ============================================================
# NO FILE MESSAGE
# ============================================================

else:

    st.info(
        "👆 Upload a CSV file above to start "
        "analysing your data."
    )


# ============================================================
# FOOTER
# ============================================================

st.divider()

st.caption(
    "Excel Report Generator | Python Data Analysis Internship Project"
)